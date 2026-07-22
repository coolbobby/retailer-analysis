"""
Core data-loading and workbook-building logic for the Retailer Analysis app.

This module is UI-agnostic: it accepts file paths or file-like objects (as
returned by Streamlit's file_uploader), and returns plain Python data
structures plus a ready-to-download Excel workbook. app.py wraps this in the
Streamlit UI.
"""

import io
import os
import csv
import calendar
import datetime
from collections import Counter, defaultdict

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule


# ---------------------------------------------------------------------------
# Loading
# ---------------------------------------------------------------------------

def _read_csv_rows(file_or_path):
    """Read a CSV from a path or an uploaded file-like object, returning
    (fieldnames, list-of-dict rows)."""
    if hasattr(file_or_path, "read"):
        raw = file_or_path.read()
        if isinstance(raw, bytes):
            raw = raw.decode("utf-8-sig")
        text_stream = io.StringIO(raw)
    else:
        text_stream = open(file_or_path, newline="", encoding="utf-8-sig")
    try:
        reader = csv.DictReader(text_stream)
        fields = reader.fieldnames
        rows = list(reader)
    finally:
        text_stream.close()
    return fields, rows


def load_pos_file(file_or_path):
    fields, rows = _read_csv_rows(file_or_path)
    active_rows = [r for r in rows if (r.get("state") or "").strip().upper() == "ACTIVE"]
    return {"fields": fields, "rows": rows, "active_rows": active_rows}


def load_bcd_file(file_or_path):
    fields, rows = _read_csv_rows(file_or_path)
    return {"fields": fields, "rows": rows}


def load_dbn_file(file_or_path):
    if hasattr(file_or_path, "read"):
        data = file_or_path.read()
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    else:
        wb = openpyxl.load_workbook(file_or_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header_raw = [c.value for c in ws[1]]
    fields = [h if h not in (None, "") else "CATEGORY" for h in header_raw]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        row_dict = {fields[i]: ("" if r[i] is None else r[i]) for i in range(len(fields))}
        if row_dict.get("USERNAME") not in (None, ""):
            row_dict["USERNAME"] = str(row_dict["USERNAME"]).strip()
        rows.append(row_dict)
    wb.close()
    by_username = {}
    for r in rows:
        u = str(r.get("USERNAME", "")).strip()
        if u and u not in by_username:
            by_username[u] = r
    return {"fields": fields, "rows": rows, "by_username": by_username}


def build_daywise_pivot(bcd_rows):
    """username -> {day -> count}, plus every day of that month (1st through
    the last day), not just the dates that happen to appear in the data."""
    pivot = defaultdict(lambda: defaultdict(int))
    for r in bcd_rows:
        u = (r.get("DE_USERNAME") or "").strip()
        d = (r.get("HLR_FINAL_ACT_DATE") or "").strip().split(" ")[0]
        if u and d:
            pivot[u][d] += 1

    dates_present = sorted({d for u in pivot for d in pivot[u]})
    if dates_present:
        year, month = (int(x) for x in dates_present[0].split("-")[:2])
    else:
        today = datetime.date.today()
        year, month = today.year, today.month
    days_in_month = calendar.monthrange(year, month)[1]
    day_values = [f"{year:04d}-{month:02d}-{d:02d}" for d in range(1, days_in_month + 1)]
    return pivot, day_values


# ---------------------------------------------------------------------------
# Username lookup (used directly by the Streamlit "Username Lookup" tab,
# without needing to build or re-read the Excel file)
# ---------------------------------------------------------------------------

POS_PROFILE_FIELDS = [
    ("Username", "username"), ("Retailer Name", "name"), ("POS Name", "pos_name_ss"),
    ("POS Owner Name", "pos_owner_name"), ("CSC Code", "csccode"),
    ("POS Unique Code", "pos_unique_code"), ("C-Topup No", "ctopupno"),
    ("NESL Done", "nesl_done"), ("Latitude", "latitude"), ("Longitude", "longitude"),
    ("State", "state"), ("Franchisee / CSC", "attached_to"),
]

DBN_PROFILE_FIELDS = [
    ("DBN Store Code", "Store Code"), ("DBN Locality", "Locality"),
    ("DBN Sub Locality", "Sub Locality"), ("District", "DISTRICT"), ("Division", "DIVISION"),
    ("Contact Number", "Contact number"), ("DBN Status", "STATUS"), ("Pincode", "PINCODE"),
    ("BTS ID", "BTS ID"), ("BTS Name", "BTS NAME"), ("BTS Latitude", "BTS LATITUDE"),
    ("BTS Longitude", "BTS LONGITUDE"), ("Category", "CATEGORY"),
]


def get_username_profile(username, pos_active_rows, dbn_by_username, pivot, day_values):
    """Return a dict describing everything the Username Lookup screen needs
    for one username, or None if the username isn't an active POS."""
    username = (username or "").strip()
    pos_row = next((r for r in pos_active_rows if (r.get("username") or "").strip() == username), None)
    if pos_row is None:
        return None

    dbn_row = dbn_by_username.get(username, {})
    day_counts = [pivot.get(username, {}).get(d, 0) for d in day_values]
    total = sum(day_counts)

    pos_profile = [(label, pos_row.get(field, "")) for label, field in POS_PROFILE_FIELDS]
    dbn_profile = [(label, dbn_row.get(field, "")) for label, field in DBN_PROFILE_FIELDS]

    return {
        "username": username,
        "pos_profile": pos_profile,
        "dbn_profile": dbn_profile,
        "has_dbn_record": bool(dbn_row),
        "sims_sold": total,
        "day_values": day_values,
        "day_counts": day_counts,
    }


# ---------------------------------------------------------------------------
# Workbook building (mirrors build_xlsx.py, adapted to accept already-loaded
# in-memory data instead of hardcoded file paths)
# ---------------------------------------------------------------------------

def build_workbook(pos_data, bcd_data, dbn_data, progress_callback=None):
    """Build the full multi-sheet workbook. Returns (BytesIO, stats-dict).
    progress_callback(str) is called with short status messages if given."""

    def note(msg):
        if progress_callback:
            progress_callback(msg)

    fields = pos_data["fields"]
    rows = pos_data["rows"]
    active_rows = pos_data["active_rows"]
    bcd_fields = bcd_data["fields"]
    bcd_rows = bcd_data["rows"]
    dbn_fields = dbn_data["fields"]
    dbn_rows = dbn_data["rows"]
    dbn_by_username = dbn_data["by_username"]

    missing = [f for f in ["username", "attached_to", "state"] if f not in fields]
    if missing:
        raise ValueError(f"POS file is missing required column(s): {missing}")
    if "DE_USERNAME" not in bcd_fields or "HLR_FINAL_ACT_DATE" not in bcd_fields:
        raise ValueError("BCD file must contain DE_USERNAME and HLR_FINAL_ACT_DATE columns")
    if "USERNAME" not in dbn_fields:
        raise ValueError("Retailer_DBN_Division_Wise file must contain a USERNAME column")

    note("Aggregating activations...")
    pivot, day_values = build_daywise_pivot(bcd_rows)
    pivot_usernames = sorted(pivot.keys())

    wb = Workbook()
    wb._named_styles[0].font = Font(name="Arial")

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    normal_font = Font(name="Arial")

    display_fields = ["Franchisee_CSC" if f == "attached_to" else f for f in fields]

    # --- POS Details ---
    note("Building POS Details sheet...")
    ws_all = wb.active
    ws_all.title = "POS Details"
    ws_all.append(display_fields)
    for cell in ws_all[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for r in rows:
        ws_all.append([r.get(f, "") for f in fields])
    ws_all.freeze_panes = "A2"
    for i, f in enumerate(fields, 1):
        ws_all.column_dimensions[get_column_letter(i)].width = max(12, min(30, len(f) + 4))

    # --- Active_POS ---
    note("Building Active_POS sheet...")
    ws_active = wb.create_sheet("Active_POS")
    ws_active.append(display_fields)
    for cell in ws_active[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for r in active_rows:
        ws_active.append([r.get(f, "") for f in fields])
    ws_active.freeze_panes = "A2"
    for i, f in enumerate(fields, 1):
        ws_active.column_dimensions[get_column_letter(i)].width = max(12, min(30, len(f) + 4))
    active_data_rows = len(active_rows)

    # --- Franchisee_CSC_Wise ---
    note("Building Franchisee_CSC_Wise summary...")
    ws_grp = wb.create_sheet("Franchisee_CSC_Wise")
    ws_grp.append(["Franchisee_CSC", "Count of POS"])
    for cell in ws_grp[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    unique_vals = sorted(set(r.get("attached_to", "") for r in active_rows), key=lambda x: (x is None, x))
    active_attached_counts = Counter(r.get("attached_to", "") for r in active_rows)
    for i, val in enumerate(unique_vals, start=2):
        ws_grp.cell(row=i, column=1, value=val).font = normal_font
        ws_grp.cell(row=i, column=2, value=active_attached_counts.get(val, 0)).font = normal_font
    total_row = len(unique_vals) + 2
    ws_grp.cell(row=total_row, column=1, value="Total").font = Font(name="Arial", bold=True)
    ws_grp.cell(row=total_row, column=2, value=sum(active_attached_counts.values())).font = Font(name="Arial", bold=True)
    ws_grp.column_dimensions["A"].width = 24
    ws_grp.column_dimensions["B"].width = 16
    ws_grp.freeze_panes = "A2"

    # --- Monthly_Activations ---
    note("Building Monthly_Activations sheet...")
    ws_bcd = wb.create_sheet("Monthly_Activations")
    ws_bcd.append(bcd_fields)
    for cell in ws_bcd[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for r in bcd_rows:
        ws_bcd.append([r.get(f, "") for f in bcd_fields])
    ws_bcd.freeze_panes = "A2"
    for i, f in enumerate(bcd_fields, 1):
        ws_bcd.column_dimensions[get_column_letter(i)].width = max(12, min(30, len(f) + 4))

    # --- Daywise_Pivot ---
    note("Building Daywise_Pivot sheet...")
    ws_pivot = wb.create_sheet("Daywise_Pivot")
    pivot_header = ["DE_USERNAME", "Total"] + day_values
    ws_pivot.append(pivot_header)
    for cell in ws_pivot[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for u in pivot_usernames:
        day_counts = [pivot[u].get(d, 0) for d in day_values]
        ws_pivot.append([u, sum(day_counts)] + day_counts)
    ws_pivot.freeze_panes = "A2"
    for i, f in enumerate(pivot_header, 1):
        ws_pivot.column_dimensions[get_column_letter(i)].width = max(12, min(30, len(f) + 4))
    pivot_data_rows = len(pivot_usernames)
    pivot_day_col_letters = [get_column_letter(3 + i) for i in range(len(day_values))]

    # --- Retailer_DBN_Division_Wise ---
    note("Building Retailer_DBN_Division_Wise sheet...")
    ws_dbn = wb.create_sheet("Retailer_DBN_Division_Wise")
    ws_dbn.append(dbn_fields)
    for cell in ws_dbn[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for r in dbn_rows:
        ws_dbn.append([r.get(f, "") for f in dbn_fields])
    ws_dbn.freeze_panes = "A2"
    for i, f in enumerate(dbn_fields, 1):
        ws_dbn.column_dimensions[get_column_letter(i)].width = max(12, min(30, len(f) + 4))

    # --- Locality_Wise_Sales ---
    note("Building Locality_Wise_Sales sheet...")
    locality_sales = defaultdict(int)
    locality_pos_count = defaultdict(int)
    for r in dbn_rows:
        loc = (r.get("Locality") or "").strip() or "UNSPECIFIED"
        u = str(r.get("USERNAME", "")).strip()
        sims = sum(pivot.get(u, {}).values())
        locality_sales[loc] += sims
        locality_pos_count[loc] += 1

    ws_loc = wb.create_sheet("Locality_Wise_Sales")
    ws_loc.append(["DBN Locality", "POS Count", "Total Sims Sold (This Month)"])
    for cell in ws_loc[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    ws_loc.cell(row=1, column=1).comment = Comment(
        "Sums each DBN Locality's retailers' Sims_Sold this month (matched via "
        "Retailer_DBN_Division_Wise USERNAME against Daywise_Pivot). Values, not "
        "live formulas -- re-run the import to refresh.",
        "Claude",
    )
    for loc in sorted(locality_sales, key=lambda x: -locality_sales[x]):
        ws_loc.append([loc, locality_pos_count[loc], locality_sales[loc]])
    ws_loc.freeze_panes = "A2"
    ws_loc.column_dimensions["A"].width = 26
    ws_loc.column_dimensions["B"].width = 14
    ws_loc.column_dimensions["C"].width = 26

    # --- POS_Sims_Ranking ---
    note("Building POS_Sims_Ranking sheet...")
    active_rows_by_username = {}
    for r in active_rows:
        u = (r.get("username") or "").strip()
        if u and u not in active_rows_by_username:
            active_rows_by_username[u] = r

    active_username_sims = []
    for r in active_rows:
        u = (r.get("username") or "").strip()
        active_username_sims.append((u, sum(pivot.get(u, {}).values())))

    _sorted_for_rank = sorted(active_username_sims, key=lambda x: -x[1])
    username_rank = {}
    _prev_sims, _prev_rank = None, 0
    for _idx, (_u, _sims) in enumerate(_sorted_for_rank, start=1):
        if _sims == _prev_sims:
            _rank = _prev_rank
        else:
            _rank = _idx
            _prev_rank, _prev_sims = _rank, _sims
        username_rank[_u] = _rank

    RANK_DETAIL_FIELDS = [
        ("Name", "name"), ("POS Name", "pos_name_ss"), ("POS Owner Name", "pos_owner_name"),
        ("Franchisee_CSC", "attached_to"), ("CSC Code", "csccode"),
    ]
    ws_rank = wb.create_sheet("POS_Sims_Ranking")
    ws_rank_header = ["Username"] + [label for label, _ in RANK_DETAIL_FIELDS] + ["DBN Locality", "Sims_Sold", "Rank"]
    ws_rank.append(ws_rank_header)
    for cell in ws_rank[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    ws_rank.cell(row=1, column=1).comment = Comment(
        "Every ACTIVE POS username ranked by total Sims_Sold this month (1 = highest), with "
        "POS and franchisee details for reference. Ties share the same rank. Values, not "
        "live formulas -- re-run the import to refresh.",
        "Claude",
    )
    for u, sims in sorted(active_username_sims, key=lambda x: username_rank[x[0]]):
        pos_row = active_rows_by_username.get(u, {})
        dbn_row = dbn_by_username.get(u, {})
        detail_values = [pos_row.get(field, "") for _, field in RANK_DETAIL_FIELDS]
        ws_rank.append([u] + detail_values + [dbn_row.get("Locality", ""), sims, username_rank[u]])
    ws_rank.freeze_panes = "A2"
    for i, f in enumerate(ws_rank_header, 1):
        ws_rank.column_dimensions[get_column_letter(i)].width = max(12, min(28, len(f) + 4))
    rank_col_letter = get_column_letter(ws_rank_header.index("Rank") + 1)

    # --- Username_Lookup dashboard ---
    note("Building Username_Lookup dashboard...")
    _build_username_lookup_sheet(
        wb, fields, dbn_fields, active_rows, active_data_rows,
        day_values, pivot_day_col_letters, rank_col_letter,
    )

    # --- One sheet per Franchisee_CSC ---
    note("Building per-Franchisee_CSC sheets...")
    _build_per_franchise_sheets(
        wb, fields, active_rows, unique_vals, ws_grp, dbn_fields, dbn_by_username,
        pivot, day_values, header_font, header_fill, normal_font,
    )

    note("Finalizing workbook...")
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    stats = {
        "total_pos_rows": len(rows),
        "active_pos_rows": active_data_rows,
        "unique_franchisees": len(unique_vals),
        "bcd_rows": len(bcd_rows),
        "dbn_rows": len(dbn_rows),
        "dbn_matched_usernames": len(dbn_by_username),
        "sheets": len(wb.sheetnames),
    }
    return buf, stats


def _build_username_lookup_sheet(wb, fields, dbn_fields, active_rows, active_data_rows,
                                  day_values, pivot_day_col_letters, rank_col_letter):
    ws_lk = wb.create_sheet("Username_Lookup")

    NAVY = "1F3864"
    ACCENT = "2E75B6"
    LIGHT_BLUE = "DCE6F1"
    LIGHT_GREY = "F2F2F2"
    GOLD = "FFD966"
    TITLE_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=16)
    SUBTITLE_FONT = Font(name="Arial", italic=True, color="D9D9D9", size=10)
    SECTION_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    LABEL_FONT = Font(name="Arial", bold=True, color="1F3864", size=10)
    VALUE_FONT = Font(name="Arial", color="000000", size=10)
    INPUT_FONT = Font(name="Arial", bold=True, color="1F3864", size=12)
    thin = Side(style="thin", color="BFBFBF")
    box_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def section_bar(row, text, span, start_col):
        ws_lk.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col + span - 1)
        c = ws_lk.cell(row=row, column=start_col, value=text)
        c.font = SECTION_FONT
        c.fill = PatternFill(start_color=ACCENT, end_color=ACCENT, fill_type="solid")
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws_lk.row_dimensions[row].height = 20

    def label_value(row, label, formula, start_col):
        lc = ws_lk.cell(row=row, column=start_col, value=label)
        lc.font = LABEL_FONT
        lc.fill = PatternFill(start_color=LIGHT_GREY, end_color=LIGHT_GREY, fill_type="solid")
        lc.border = box_border
        lc.alignment = Alignment(vertical="center", indent=1)
        vc = ws_lk.cell(row=row, column=start_col + 1, value=formula)
        vc.font = VALUE_FONT
        vc.border = box_border
        vc.alignment = Alignment(vertical="center", indent=1)

    username_col_letter_active = get_column_letter(fields.index("username") + 1)

    def pos_lookup(field_name):
        col = get_column_letter(fields.index(field_name) + 1)
        idx = (f"INDEX(Active_POS!{col}:{col},"
               f"MATCH($C$4,Active_POS!${username_col_letter_active}:${username_col_letter_active},0))")
        return f'=IFERROR(IF({idx}="","",{idx}),"Not Found")'

    def dbn_lookup(field_name):
        col = get_column_letter(dbn_fields.index(field_name) + 1)
        idx = (f"INDEX(Retailer_DBN_Division_Wise!{col}:{col},"
               f"MATCH($C$4,Retailer_DBN_Division_Wise!$A:$A,0))")
        return f'=IFERROR(IF({idx}="","",{idx}),"No DBN Record")'

    ws_lk.merge_cells("A1:F1")
    t = ws_lk["A1"]
    t.value = "RETAILER USERNAME LOOKUP & MONTHLY ACTIVATION ANALYSIS"
    t.font = TITLE_FONT
    t.fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws_lk.row_dimensions[1].height = 30

    ws_lk.merge_cells("A2:F2")
    st_ = ws_lk["A2"]
    st_.value = "Enter a username below to pull that retailer's full profile, division details, and this month's SIM activation breakdown."
    st_.font = SUBTITLE_FONT
    st_.fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
    st_.alignment = Alignment(horizontal="center", vertical="center")
    ws_lk.row_dimensions[2].height = 18

    ws_lk["B4"] = "TYPE USERNAME ▶"
    ws_lk["B4"].font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    ws_lk["B4"].fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
    ws_lk["B4"].alignment = Alignment(horizontal="right", vertical="center", indent=1)
    ws_lk.row_dimensions[4].height = 24

    ws_lk.merge_cells("C4:D4")
    input_cell = ws_lk["C4"]
    input_cell.value = active_rows[0].get("username", "") if active_rows else ""
    input_cell.font = INPUT_FONT
    input_cell.fill = PatternFill(start_color=GOLD, end_color=GOLD, fill_type="solid")
    input_cell.alignment = Alignment(horizontal="center", vertical="center")
    input_cell.border = box_border

    # Free-text entry -- no dropdown. A prompt still shows on selecting the cell
    # to guide the user, but any username can be typed directly.
    dv = DataValidation(type="textLength", operator="greaterThan", formula1="0", allow_blank=True)
    dv.prompt = "Type the retailer username (mobile number) and press Enter"
    dv.promptTitle = "Enter Username"
    dv.error = "Please enter a username."
    dv.errorTitle = "Username required"
    ws_lk.add_data_validation(dv)
    dv.add(input_cell)

    ws_lk.merge_cells("E4:F4")
    status_cell = ws_lk["E4"]
    status_cell.value = (
        f'=IF($C$4="","Enter a username above",'
        f'IF(ISNA(MATCH($C$4,Active_POS!${username_col_letter_active}:${username_col_letter_active},0)),'
        f'"Not found in Active_POS","Match found"))'
    )
    status_cell.font = Font(name="Arial", bold=True, size=10)
    status_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_lk.conditional_formatting.add(
        "E4:F4",
        CellIsRule(operator="equal", formula=['"Not found in Active_POS"'],
                   fill=PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")),
    )
    ws_lk.conditional_formatting.add(
        "E4:F4",
        CellIsRule(operator="equal", formula=['"Match found"'],
                   fill=PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")),
    )

    row = 6
    section_bar(row, "POS PROFILE", 2, 1)
    section_bar(row, "FRANCHISEE / DIVISION", 2, 4)
    row += 1
    pos_start_row = row
    pos_field_labels = [
        ("Username", "username"), ("Retailer Name", "name"), ("POS Name", "pos_name_ss"),
        ("POS Owner Name", "pos_owner_name"), ("CSC Code", "csccode"),
        ("POS Unique Code", "pos_unique_code"), ("C-Topup No", "ctopupno"),
        ("NESL Done", "nesl_done"), ("Latitude", "latitude"), ("Longitude", "longitude"),
        ("State", "state"),
    ]
    franchisee_field_labels = [("Franchisee / CSC", "attached_to")]
    dbn_field_labels = [
        ("DBN Store Code", "Store Code"), ("DBN Locality", "Locality"), ("DBN Sub Locality", "Sub Locality"),
        ("District", "DISTRICT"), ("Division", "DIVISION"), ("Contact Number", "Contact number"),
        ("DBN Status", "STATUS"), ("Pincode", "PINCODE"), ("BTS ID", "BTS ID"),
        ("BTS Name", "BTS NAME"), ("BTS Latitude", "BTS LATITUDE"), ("BTS Longitude", "BTS LONGITUDE"),
        ("Category", "CATEGORY"),
    ]
    for label, field in pos_field_labels:
        label_value(row, label, pos_lookup(field), 1)
        row += 1
    pos_end_row = row - 1

    row2 = pos_start_row
    locality_value_row = None
    for label, field in franchisee_field_labels:
        label_value(row2, label, pos_lookup(field), 4)
        row2 += 1
    for label, field in dbn_field_labels:
        label_value(row2, label, dbn_lookup(field), 4)
        if field == "Locality":
            locality_value_row = row2
        row2 += 1
    franchise_end_row = row2 - 1
    locality_value_cell = f"E{locality_value_row}"

    next_row = max(pos_end_row, franchise_end_row) + 2
    section_bar(next_row, "MONTHLY ACTIVATION ANALYSIS (SIMS SOLD)", 6, 1)
    next_row += 1

    label_value(next_row, "Total Sims Sold (This Month)",
                "=IFERROR(INDEX(Daywise_Pivot!$B:$B,MATCH($C$4,Daywise_Pivot!$A:$A,0)),0)", 1)
    ws_lk.cell(row=next_row, column=2).font = Font(name="Arial", bold=True, color="1F3864", size=14)
    ws_lk.cell(row=next_row, column=2).fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
    next_row += 2

    table_header_row = next_row
    ws_lk.cell(row=table_header_row, column=1, value="Date").font = Font(name="Arial", bold=True, color="FFFFFF")
    ws_lk.cell(row=table_header_row, column=2, value="Activations").font = Font(name="Arial", bold=True, color="FFFFFF")
    for c in (1, 2):
        ws_lk.cell(row=table_header_row, column=c).fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
        ws_lk.cell(row=table_header_row, column=c).alignment = Alignment(horizontal="center")

    day_table_start = table_header_row + 1
    for d_offset, day_val in enumerate(day_values):
        r = day_table_start + d_offset
        dc = ws_lk.cell(row=r, column=1, value=day_val)
        dc.font = VALUE_FONT
        dc.border = box_border
        dc.alignment = Alignment(horizontal="center")
        day_col_letter = pivot_day_col_letters[d_offset]
        vc = ws_lk.cell(
            row=r, column=2,
            value=f"=IFERROR(INDEX(Daywise_Pivot!{day_col_letter}:{day_col_letter},"
                  f"MATCH($C$4,Daywise_Pivot!$A:$A,0)),0)",
        )
        vc.font = VALUE_FONT
        vc.border = box_border
        vc.alignment = Alignment(horizontal="center")
        if (d_offset % 2) == 1:
            dc.fill = PatternFill(start_color=LIGHT_GREY, end_color=LIGHT_GREY, fill_type="solid")
            vc.fill = PatternFill(start_color=LIGHT_GREY, end_color=LIGHT_GREY, fill_type="solid")
    day_table_end = day_table_start + len(day_values) - 1

    # --- Performance Insights panel (locality sales + ranking), beside the day table ---
    section_bar(table_header_row, "PERFORMANCE INSIGHTS", 3, 4)
    insight_row = table_header_row + 1

    label_value(insight_row, "DBN Locality", f"={locality_value_cell}", 4)
    insight_row += 1

    label_value(
        insight_row, "Locality POS Count",
        f'=IFERROR(INDEX(Locality_Wise_Sales!$B:$B,MATCH({locality_value_cell},Locality_Wise_Sales!$A:$A,0)),"—")',
        4,
    )
    insight_row += 1

    label_value(
        insight_row, "Locality Total Sims Sold",
        f'=IFERROR(INDEX(Locality_Wise_Sales!$C:$C,MATCH({locality_value_cell},Locality_Wise_Sales!$A:$A,0)),"—")',
        4,
    )
    ws_lk.cell(row=insight_row, column=5).font = Font(name="Arial", bold=True, color="1F3864", size=11)
    insight_row += 2

    label_value(
        insight_row, "Username Rank (by Sims Sold)",
        f'=IFERROR(INDEX(POS_Sims_Ranking!${rank_col_letter}:${rank_col_letter},MATCH($C$4,POS_Sims_Ranking!$A:$A,0)),"N/A")',
        4,
    )
    ws_lk.cell(row=insight_row, column=5).font = Font(name="Arial", bold=True, color="1F3864", size=14)
    ws_lk.cell(row=insight_row, column=5).fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
    insight_row += 1

    label_value(
        insight_row, "Out Of (Active Usernames)",
        f"=COUNTA(Active_POS!${username_col_letter_active}$2:${username_col_letter_active}$100000)",
        4,
    )

    ws_lk.column_dimensions["A"].width = 22
    ws_lk.column_dimensions["B"].width = 16
    ws_lk.column_dimensions["C"].width = 20
    ws_lk.column_dimensions["D"].width = 22
    ws_lk.column_dimensions["E"].width = 20
    ws_lk.column_dimensions["F"].width = 16
    ws_lk.sheet_view.showGridLines = False
    ws_lk.freeze_panes = None

    wb.move_sheet("Username_Lookup", offset=-(len(wb.sheetnames) - 1))


def _build_per_franchise_sheets(wb, fields, active_rows, unique_vals, ws_grp, dbn_fields,
                                 dbn_by_username, pivot, day_values, header_font, header_fill, normal_font):
    PER_VAL_FIELDS = [
        "username", "ctopupno", "name", "pos_name_ss", "pos_owner_name",
        "nesl_done", "pos_unique_code", "csccode", "latitude", "longitude",
    ]
    SIMS_SOLD_COL = "Sims_Sold"
    DBN_HEADER = [f"DBN_{f}" for f in dbn_fields]
    PER_VAL_HEADER = PER_VAL_FIELDS + DBN_HEADER + [SIMS_SOLD_COL] + day_values
    dbn_col_start_idx = len(PER_VAL_FIELDS) + 1
    sims_sold_col_idx = len(PER_VAL_FIELDS) + len(DBN_HEADER) + 1
    day_col_start_idx = sims_sold_col_idx + 1

    latlong_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    dbn_fill = PatternFill(start_color="D9D2E9", end_color="D9D2E9", fill_type="solid")
    sims_sold_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    daywise_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    lat_idx_pv = PER_VAL_FIELDS.index("latitude")
    lon_idx_pv = PER_VAL_FIELDS.index("longitude")

    used_names = set(wb.sheetnames)

    for i, val in enumerate(unique_vals, start=2):
        safe_name = (val or "BLANK").strip()
        for ch in [":", "\\", "/", "?", "*", "[", "]"]:
            safe_name = safe_name.replace(ch, "_")
        safe_name = safe_name[:31]
        base_name = safe_name
        suffix = 1
        while safe_name in used_names:
            suffix += 1
            cut = 31 - len(f"_{suffix}")
            safe_name = f"{base_name[:cut]}_{suffix}"
        used_names.add(safe_name)

        ws_val = wb.create_sheet(safe_name)
        ws_val.append(PER_VAL_HEADER)
        for cell in ws_val[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        for d_offset in range(len(DBN_HEADER)):
            ws_val.cell(row=1, column=dbn_col_start_idx + d_offset).fill = PatternFill(
                start_color="674EA7", end_color="674EA7", fill_type="solid")
        ws_val.cell(row=1, column=sims_sold_col_idx).fill = PatternFill(
            start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        for d_offset in range(len(day_values)):
            ws_val.cell(row=1, column=day_col_start_idx + d_offset).fill = PatternFill(
                start_color="C55A11", end_color="C55A11", fill_type="solid")

        val_rows = [r for r in active_rows if r.get("attached_to", "") == val]
        for r in val_rows:
            u = (r.get("username") or "").strip()
            day_counts = [pivot[u].get(d, 0) for d in day_values]
            total = sum(day_counts)
            dbn_row = dbn_by_username.get(u, {})
            dbn_values = [dbn_row.get(f, "") for f in dbn_fields]
            ws_val.append([r.get(f, "") for f in PER_VAL_FIELDS] + dbn_values + [total] + day_counts)
        for row in ws_val.iter_rows(min_row=2):
            for cell in row:
                cell.font = normal_font
            lat_val = row[lat_idx_pv].value
            lon_val = row[lon_idx_pv].value
            has_latlong = bool(str(lat_val).strip()) and bool(str(lon_val).strip())
            if has_latlong:
                for cell in row[:len(PER_VAL_FIELDS)]:
                    cell.fill = latlong_fill
            for cell in row[dbn_col_start_idx - 1:sims_sold_col_idx - 1]:
                cell.fill = dbn_fill
            row[sims_sold_col_idx - 1].fill = sims_sold_fill
            for cell in row[day_col_start_idx - 1:]:
                cell.fill = daywise_fill
        ws_val.freeze_panes = "A2"
        for c, f in enumerate(PER_VAL_HEADER, 1):
            ws_val.column_dimensions[get_column_letter(c)].width = max(12, min(30, len(f) + 4))

        grp_cell = ws_grp.cell(row=i, column=1)
        grp_cell.hyperlink = f"#'{safe_name}'!A1"
        grp_cell.font = Font(name="Arial", color="0563C1", underline="single")
